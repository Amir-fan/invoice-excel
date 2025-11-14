import os
import tempfile
from typing import Optional
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from dotenv import load_dotenv

# Load environment variables from .env file (only in local development)
# Vercel doesn't support .env files - use environment variables from dashboard instead
try:
    load_dotenv()
except Exception:
    pass  # Ignore .env loading errors in production

# Lazy imports to prevent crashes during module load on Vercel
# Import only when functions are actually called
try:
    from ai import (
        extract_invoice_data_from_image,
        extract_invoice_data_from_text,
        extract_invoice_data_from_pdf_text_with_lines,
    )
    from mapping import create_invoice_rows, ARABIC_HEADERS
    from utils import (
        pdf_to_image,
        image_to_bytes,
        create_temp_file,
        cleanup_temp_file,
        is_pdf_file,
        is_image_file,
        check_pdf_dependencies,
        get_pdf_installation_instructions,
        extract_pdf_text_lines,
    )
except ImportError as e:
    # If imports fail, log but don't crash - handlers will show error
    import sys
    print(f"WARNING: Import failed: {e}", file=sys.stderr)
    # Set to None so we can check later
    extract_invoice_data_from_image = None
    extract_invoice_data_from_text = None
    extract_invoice_data_from_pdf_text_with_lines = None
    create_invoice_rows = None
    ARABIC_HEADERS = None

app = FastAPI(title="Invoice2Excel Web", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with specific origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add logging middleware (only in debug mode)
DEBUG = os.getenv("DEBUG", "false").lower() == "true"

if DEBUG:
    @app.middleware("http")
    async def log_requests(request, call_next):
        print(f"DEBUG: {request.method} {request.url.path}")
        response = await call_next(request)
        print(f"DEBUG: Response status: {response.status_code}")
        return response

# Templates - use absolute path for Vercel compatibility
from pathlib import Path

# Get the directory where app.py is located
BASE_DIR = Path(__file__).parent
TEMPLATES_DIR = BASE_DIR / "templates"

# Initialize templates - handle case where directory might not exist
try:
    templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
except Exception as e:
    print(f"Warning: Could not initialize Jinja2Templates: {e}")
    templates = None

@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Serve the main HTML page."""
    try:
        html_path = TEMPLATES_DIR / "index.html"
        with open(html_path, "r", encoding="utf-8") as f:
            html_content = f.read()
        return HTMLResponse(content=html_content)
    except Exception as e:
        # Fallback if template file not found
        print(f"ERROR reading template: {e}")
        return HTMLResponse(
            content=f"<html><body><h1>Error</h1><p>Template file not found: {e}</p></body></html>",
            status_code=500
        )

@app.get("/health")
async def health_check():
    """Health check endpoint with dependency status."""
    pdf_deps_ok = check_pdf_dependencies()
    instructions = get_pdf_installation_instructions()
    
    # Check if OpenAI API key is configured (don't validate it to avoid slow health checks)
    try:
        from ai import get_openai_client
        api_key = os.getenv("OPENAI_API_KEY")
        api_key_configured = api_key is not None and len(api_key) > 0
        api_key_status = "configured" if api_key_configured else "not configured"
    except Exception:
        api_key_status = "unknown"
    
    return {
        "status": "healthy", 
        "service": "Invoice2Excel Web",
        "pdf_processing": "available" if pdf_deps_ok else "unavailable",
        "pdf_instructions": instructions,
        "openai_api_key": api_key_status,
        "supported_formats": ["PDF", "JPG", "JPEG", "PNG"] if pdf_deps_ok else ["JPG", "JPEG", "PNG"]
    }

def _process_single_invoice(file: UploadFile, temp_file_path: str) -> list:
    """
    Process a single invoice file and return rows.
    Helper function to process one invoice file.
    """
    # Check if imports succeeded
    if extract_invoice_data_from_image is None:
        raise HTTPException(status_code=500, detail="Application not fully initialized - imports failed")
    
    if DEBUG:
        print(f"DEBUG: Processing file: {file.filename}")
    
    # Extract invoice data
    invoice_data = None
    
    if is_pdf_file(file.filename):
        if DEBUG:
            print(f"DEBUG: Processing PDF file: {file.filename}")

        # First, try text-based extraction: use PyMuPDF to get the exact PDF text,
        # then let GPT-4o parse that text into structured JSON, and finally
        # align each item description to the exact PDF line that contains its numbers.
        try:
            lines = extract_pdf_text_lines(temp_file_path)
            if lines and extract_invoice_data_from_pdf_text_with_lines is not None:
                if DEBUG:
                    print(f"DEBUG: Extracted {len(lines)} text line(s) from PDF")
                    print("DEBUG: Calling extract_invoice_data_from_pdf_text_with_lines (GPT over text + alignment)...")
                text = "\n".join(lines)
                invoice_data = extract_invoice_data_from_pdf_text_with_lines(text, lines)
                if invoice_data and DEBUG:
                    print("DEBUG: GPT text-based extraction + alignment succeeded")
        except Exception as e:
            print(f"WARNING: Text-based GPT extraction failed for {file.filename}: {e}")

        # If text-based extraction failed, fall back to image + vision
        if not invoice_data:
            if DEBUG:
                print("DEBUG: Falling back to PDF-to-image + AI vision pipeline")
            image = pdf_to_image(temp_file_path)
            if image:
                if DEBUG:
                    print(f"DEBUG: PDF converted to image: {image.size}")
                    print("DEBUG: Calling AI image extraction...")
                image_bytes = image_to_bytes(image)
                invoice_data = extract_invoice_data_from_image(image_bytes)
                if DEBUG:
                    print("DEBUG: AI image extraction completed")
            else:
                if DEBUG:
                    print("DEBUG: PDF to image conversion failed")
    elif is_image_file(file.filename):
        # Read image file directly
        with open(temp_file_path, "rb") as f:
            image_bytes = f.read()
        invoice_data = extract_invoice_data_from_image(image_bytes)
    
    if not invoice_data:
        print(f"WARNING: Failed to extract data from {file.filename}")
        return []
    
    # Create rows from invoice data (one row per line item)
    rows = create_invoice_rows(invoice_data)
    if DEBUG:
        print(f"DEBUG: Created {len(rows)} row(s) from invoice data")
    
    return rows


@app.post("/upload")
async def upload_files(files: list[UploadFile] = File(...)):
    """
    Process uploaded invoice files and return Excel with extracted data.
    Supports multiple files - all invoices will be combined into one Excel file.
    Supports PDF, JPG, JPEG, PNG files.
    """
    if DEBUG:
        print(f"DEBUG: Received POST request to /upload")
        print(f"DEBUG: Files parameter type: {type(files)}")
        print(f"DEBUG: Processing {len(files) if isinstance(files, list) else 1} file(s)")
    
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    
    # Handle both single file and multiple files
    if not isinstance(files, list):
        files = [files]
    
    # Validate all files
    for file in files:
        if not (is_pdf_file(file.filename) or is_image_file(file.filename)):
            raise HTTPException(
                status_code=400, 
                detail=f"Unsupported file type: {file.filename}. Please upload PDF, JPG, JPEG, or PNG files only."
            )
        
        # Check PDF dependencies if uploading a PDF
        if is_pdf_file(file.filename) and not check_pdf_dependencies():
            instructions = get_pdf_installation_instructions()
            raise HTTPException(
                status_code=400,
                detail=f"PDF processing is not available. {instructions}\n\nQuick fix: Run this command in your terminal:\npip install PyMuPDF\n\nThen restart the application."
            )
    
    # Process all files
    all_rows = []
    temp_files = []
    processed_count = 0
    failed_files = []
    
    try:
        for idx, file in enumerate(files, 1):
            temp_file_path = None
            try:
                if DEBUG:
                    print(f"DEBUG: Processing file {idx}/{len(files)}: {file.filename}")
                
                # Save uploaded file temporarily
                temp_file_path = create_temp_file(os.path.splitext(file.filename)[1])
                temp_files.append(temp_file_path)
                
                with open(temp_file_path, "wb") as buffer:
                    content = await file.read()
                    buffer.write(content)
                
                # Process the invoice
                rows = _process_single_invoice(file, temp_file_path)
                
                if rows:
                    all_rows.extend(rows)
                    processed_count += 1
                    if DEBUG:
                        print(f"DEBUG: Successfully processed {file.filename}, added {len(rows)} row(s)")
                else:
                    failed_files.append(file.filename)
                    print(f"WARNING: Failed to extract data from {file.filename}")
                    
            except Exception as e:
                print(f"ERROR: Failed to process {file.filename}: {e}")
                failed_files.append(file.filename)
                import traceback
                traceback.print_exc()
            finally:
                # Clean up temp file immediately after processing
                if temp_file_path and os.path.exists(temp_file_path):
                    cleanup_temp_file(temp_file_path)
                    if temp_file_path in temp_files:
                        temp_files.remove(temp_file_path)
        
        # If we got no data at all, create placeholder rows for failed files
        # instead of raising an error, so the user still gets an Excel file
        if not all_rows:
            if failed_files:
                if DEBUG:
                    print("DEBUG: No successful extractions, creating placeholder rows.")
                all_rows = []
                for fname in failed_files:
                    # Create an empty row with a message in the description column
                    row = {header: "" for header in ARABIC_HEADERS}
                    row["الوصف"] = f"فشل استخراج البيانات من الملف: {fname}"
                    all_rows.append(row)
            else:
                # This should not normally happen, but keep the old error as a fallback
                raise HTTPException(
                    status_code=500,
                    detail="Failed to extract data from any files (no details available)."
                )
        
        # Create Excel file with all rows
        if DEBUG:
            print(f"DEBUG: Creating Excel file with {len(all_rows)} total row(s) from {processed_count} invoice(s)...")
        excel_bytes = create_excel_file(all_rows)
        if DEBUG:
            print(f"DEBUG: Excel created: {len(excel_bytes)} bytes")
        
        # Build success message
        success_msg_parts = [f"تم معالجة {processed_count} فاتورة بنجاح"]
        if failed_files:
            success_msg_parts.append(f"({len(failed_files)} فشلت: {', '.join(failed_files)})")
        
        # Return Excel file as bytes response
        from fastapi.responses import Response
        response = Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=invoice_data_{processed_count}_invoices.xlsx"}
        )
        
        return response
        
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        # Handle other errors with detailed logging
        import traceback
        error_details = traceback.format_exc()
        error_message = str(e)
        print(f"ERROR: Error processing files: {error_message}")
        print(f"Full traceback: {error_details}")
        
        # Provide user-friendly error messages
        if "api key" in error_message.lower() or "authentication" in error_message.lower():
            user_message = "مفتاح API غير صحيح أو مفقود. يرجى التحقق من مفتاح OpenAI API."
        elif "rate limit" in error_message.lower():
            user_message = "تم تجاوز حد الاستخدام لـ OpenAI API. يرجى المحاولة لاحقاً."
        elif "quota" in error_message.lower():
            user_message = "تم تجاوز الحصة المتاحة لـ OpenAI API. يرجى التحقق من حسابك."
        else:
            user_message = f"خطأ في المعالجة: {error_message}"
        
        raise HTTPException(status_code=500, detail=user_message)
    finally:
        # Clean up any remaining temporary files
        for temp_file_path in temp_files:
            if temp_file_path and os.path.exists(temp_file_path):
                cleanup_temp_file(temp_file_path)

def create_excel_file(rows: list) -> bytes:
    """
    Create Excel file with RTL support and Arabic headers.
    Takes a list of row dictionaries (one per line item).
    Returns the file as bytes.
    """
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "الفواتير"
    
    # Set RTL (Right-to-Left) view
    ws.sheet_view.rightToLeft = True
    
    # Write headers (row 1)
    for col_idx, header in enumerate(ARABIC_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data rows (one per line item)
    for row_idx, row_data in enumerate(rows, start=2):  # Start at row 2 (after header)
        for col_idx, header in enumerate(ARABIC_HEADERS, 1):
            value = row_data.get(header, 0)

            # For numeric monetary fields we keep up to 3 decimal places (as in invoices)
            # to avoid Excel rounding 1999.998 to 2000.00 in the display.
            numeric_headers = [
                "الكمية", "سعر الوحدة", "المبلغ", "الخصم", "الاجمالي",
                "إجمالي قيمة الفاتورة", "تسلسل مصدر الدخل", "الرقم الضريبي"
            ]
            if header in numeric_headers and isinstance(value, (int, float)):
                value = round(float(value), 3)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format numeric cells
            if isinstance(value, (int, float)) and header in numeric_headers:
                # Use three decimal places to match invoice formatting
                cell.number_format = '#,##0.000'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                # Text fields - center align
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        if adjusted_width > 0:
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
